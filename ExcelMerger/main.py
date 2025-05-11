import sys
import pandas as pd
import re
from openpyxl import load_workbook
from PyQt6.QtWidgets import (QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget,QFileDialog, QMessageBox, QLabel)
from PyQt6.QtCore import Qt
from datetime import datetime


class UniversalExcelUpdater(QMainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.source_file = None
        self.target_file = None

    def init_ui(self):
        self.setWindowTitle("Универсальный обработчик Excel")
        self.setMinimumSize(500, 400)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        self.btn_load_source = QPushButton("Загрузить файл с ID")
        self.btn_load_target = QPushButton("Загрузить целевой файл")
        self.btn_process = QPushButton("Запустить обработку")

        self.lbl_source = QLabel("Файл EX1: не выбран")
        self.lbl_target = QLabel("Файл EX2: не выбран")

        layout.addWidget(self.btn_load_source)
        layout.addWidget(self.lbl_source)
        layout.addWidget(self.btn_load_target)
        layout.addWidget(self.lbl_target)
        layout.addWidget(self.btn_process)
        layout.addStretch()

        self.btn_load_source.clicked.connect(lambda: self.load_file('source'))
        self.btn_load_target.clicked.connect(lambda: self.load_file('target'))
        self.btn_process.clicked.connect(self.process_data)

    def load_file(self, file_type):
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel Files (*.xlsx *.xls)")
        if file_name:
            if file_type == 'source':
                self.source_file = file_name
                self.lbl_source.setText(f"EX1: {file_name}")
            else:
                self.target_file = file_name
                self.lbl_target.setText(f"EX2: {file_name}")

    def process_data(self):
        try:
            df_source = pd.read_excel(self.source_file, header=None, dtype=object, engine='openpyxl')
            id_mapping = {}
            for _, row in df_source.iterrows():
                passport = str(row[2]).strip().replace('.0', '') if not pd.isna(row[2]) else ''
                if passport:
                    id_mapping[passport] = row[0]

            wb = load_workbook(self.target_file)
            ws = wb.active

            added_count = 0

            for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                if self.is_header_row_openpyxl(row):
                    continue

                matched_id = None

                first_non_empty_column_idx = None
                for col_idx, cell in enumerate(row):
                    if cell.value not in [None, '']:
                        first_non_empty_column_idx = col_idx
                        break

                for col_idx, cell in enumerate(row):
                    if col_idx == first_non_empty_column_idx:
                        continue
                    if cell.value is not None:
                        passport = str(cell.value).strip().replace('.0', '')
                        if passport in id_mapping:
                            matched_id = id_mapping[passport]
                            break

                if matched_id is not None:
                    id_found_in_row = False
                    for col_idx, cell in enumerate(row):
                        if col_idx == first_non_empty_column_idx:
                            continue
                        if cell.value is not None:
                            existing_value = str(cell.value).strip().replace('.0', '')
                            match = re.match(r'^\d+', existing_value)
                            if match:
                                existing_id = match.group(0)
                                if existing_id == str(matched_id):
                                    id_found_in_row = True
                                    break

                    if id_found_in_row:
                        continue

                    last_col_idx = 1
                    for cell in row:
                        if cell.value is not None:
                            last_col_idx = cell.column
                    insert_col = last_col_idx + 1
                    ws.cell(row=row_idx, column=insert_col, value=matched_id)
                    added_count += 1

            wb.save(self.target_file)

            QMessageBox.information(self, "Успех", f"Обработка завершена!\nДобавлено ID: {added_count}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка обработки:\n{str(e)}")

    def is_header_row_openpyxl(self, row):
        try:
            first_cell = row[0].value
            if isinstance(first_cell, datetime):
                return all(cell.value is None for cell in row[1:])
            return False
        except:
            return False


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = UniversalExcelUpdater()
    window.show()
    sys.exit(app.exec())
